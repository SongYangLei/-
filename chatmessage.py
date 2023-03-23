from websocket import WebSocketApp
import json
import re
import gzip
from urllib.parse import unquote_plus
import requests
import pandas as pd
import xlwt
import xlrd
import openpyxl as op
from openpyxl import Workbook
from douyin_pb2 import PushFrame, Response, ChatMessage


def fetch_live_room_info(url):
    res = requests.get(
        url=url,
        headers={
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36",
        },
        cookies={
            "__ac_nonce": "063abcffa00ed8507d599"  # 可以是任意值
        }
    )
    data_string = re.findall(r'<script id="RENDER_DATA" type="application/json">(.*?)</script>', res.text)[0]
    data_dict = json.loads(unquote_plus(data_string))

    room_id = data_dict['app']['initialState']['roomStore']['roomInfo']['roomId']
    room_title = data_dict['app']['initialState']['roomStore']['roomInfo']["room"]['title']
    room_user_count = data_dict['app']['initialState']['roomStore']['roomInfo']["room"]['user_count_str']

    # print(room_id)
    wss_url = f'wss://webcast3-ws-web-lf.douyin.com/webcast/im/push/v2/?app_name=douyin_web&version_code=180800&webcast_sdk_version=1.3.0&update_version_code=1.3.0&compress=gzip&internal_ext=internal_src:dim|wss_push_room_id:7213507724800248632|wss_push_did:7211044968339899965|dim_log_id:2023032309594083794F4BF36D1AA0B7E9|fetch_time:1679536780419|seq:1|wss_info:0-1679536780419-0-0|wrds_kvs:WebcastRoomStatsMessage-1679536775887239771_InputPanelComponentSyncData-1679525741550524650_WebcastRoomRankMessage-1679536673800563303&cursor=t-1679536780419_r-1_d-1_u-1_h-1&host=https://live.douyin.com&aid=6383&live_id=1&did_rule=3&debug=false&maxCacheMessageNumber=20&endpoint=live_pc&support_wrds=1&im_path=/webcast/im/fetch/&user_unique_id=7211044968339899965&device_platform=web&cookie_enabled=true&screen_width=1536&screen_height=864&browser_language=zh-CN&browser_platform=Win32&browser_name=Mozilla&browser_version=5.0%20(Windows%20NT%2010.0;%20Win64;%20x64)%20AppleWebKit/537.36%20(KHTML,%20like%20Gecko)%20Chrome/111.0.0.0%20Safari/537.36%20Edg/111.0.1661.44&browser_online=true&tz_name=Asia/Shanghai&identity=audience&room_id=7213507724800248632&heartbeatDuration=0&signature=WBYe+w9+Ha48Fxo9'
    ttwid = res.cookies.get_dict()['ttwid']
    return room_id, room_title, room_user_count, wss_url, ttwid


def on_open(ws):
    print('on_open')


def on_message(ws, content):
    frame = PushFrame()
    frame.ParseFromString(content)

    # 对PushFrame的 payload 内容进行gzip解压
    origin_bytes = gzip.decompress(frame.payload)

    # 根据Response+gzip解压数据，生成数据对象
    response = Response()
    response.ParseFromString(origin_bytes)

    if response.needAck:
        s = PushFrame()
        s.payloadType = "ack"
        s.payload = response.internalExt.encode('utf-8')
        s.logId = frame.logId

        ws.send(s.SerializeToString())
    wb = op.load_workbook('3-23东方甄选10点.xlsx')
    sheet = wb['3-23东方甄选10点']
    for item in response.messagesList:
        if item.method != "WebcastChatMessage":
             continue
        message = ChatMessage()
        message.ParseFromString(item.payload)

        data = [message.user.nickName, message.user.gender, message.content, message.eventTime,message.user.shortId,message.user.id,message.user.Signature,message.user.Level,message.user.Birthday,message.user.Telephone,message.user.city]
        # print(message.user.nickName, message.user.gender, message.content, message.eventTime,message.user.shortId,message.user.id,message.user.Signature,message.user.Level,message.user.Birthday,message.user.Telephone,message.user.city)
        sheet.append(data)
    wb.save('3-23东方甄选10点.xlsx')


def on_error(ws, content):
    print(content)
    print("on_error")


def on_close(*args, **kwargs):
    print(args, kwargs)
    print("on_close")


def run():
    web_url = "https://live.douyin.com/80017709309"

    room_id, room_title, room_user_count, wss_url, ttwid = fetch_live_room_info(web_url)

    # print(room_id, room_title, room_user_count, wss_url, ttwid)

    ws = WebSocketApp(
        url=wss_url,
        header={
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36"
        },
        cookie=f"ttwid={ttwid}",
        on_open=on_open,
        on_message=on_message,
        on_error=on_error,
        on_close=on_close,
    )
    ws.run_forever()


if __name__ == '__main__':
    wb = op.Workbook()
    wb.create_sheet(index=1)
    wb.active.title = '3-23东方甄选10点'
    wb.active.cell(row=1, column=1, value='昵称')
    wb.active.cell(row=1, column=2, value='性别')
    wb.active.cell(row=1, column=3, value='聊天内容')
    wb.active.cell(row=1, column=4, value='时间戳')
    wb.active.cell(row=1, column=5, value='ID')
    wb.active.cell(row=1, column=6, value='id')
    wb.active.cell(row=1, column=7, value='签名')
    wb.active.cell(row=1, column=8, value='等级')
    wb.active.cell(row=1, column=9, value='生日')
    wb.active.cell(row=1, column=10, value='电话')
    wb.active.cell(row=1, column=11, value='城市')
    wb.save('3-23东方甄选10点.xlsx')
    run()



